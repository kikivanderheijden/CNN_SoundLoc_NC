import numpy as np
import openpyxl

def Evaluation_CalculateAveragePerformance(CNN_dict_preds, excelfile, filenames_CNN):
    
    mse = CNN_dict_preds['mse']
    cosine_distance_degrees = CNN_dict_preds['cosine_distance_degrees']
    
    # retrieve values for plot (A), you have to divide models into trained with AD loss and trained with MSE loss 
    # prepare data
    mean_mse_CNN = np.mean(mse, axis = 1)    
    mean_ad_CNN = np.mean(cosine_distance_degrees, axis = 1) # mean of the angular distance (in degrees)
    # information needed to retrieve number of model parameters
    wb = openpyxl.load_workbook(excelfile) # open excel file
    ws = wb['RawData'] # select active sheet
    modelsanalyzed = ws['A'] # get names of analyzed models from first column
    models_nrparams = ws['I']
    
    # cycle through the filenames and if it is a model of a specific type you add it to the array
    
    ## for CNN models 
    # first determine how many CNN models with MSE loss and AD loss there are to set the dimensions of the arrays
    CNN_mse = 0 # initialize counters
    CNN_ad = 0
    for w in range(len(filenames_CNN)):
        if '_MSE' in filenames_CNN[w]:
            CNN_mse = CNN_mse + 1
        elif '_AD' in filenames_CNN[w]:
            CNN_ad = CNN_ad + 1
    
    # initialize counters and arrays  
    # trained with MSE loss
    count_CNN_mse = 0 # this is the counter 
    CNN_mse_names = [] # this will be a list of model names
    CNN_mse_scoremse = np.zeros([CNN_mse,1]) # this will be an array of the MSE score 
    CNN_mse_scoread = np.zeros([CNN_mse,1]) # this will be an array of the AD score
    CNN_mse_nrparams = np.zeros([CNN_mse,1]) # thiss will be an array of the number of parameters
    # trained with AD loss
    count_CNN_ad = 0 # this is the counter 
    CNN_ad_names = [] # this will be a list of model names 
    CNN_ad_scoremse = np.zeros([CNN_ad,1]) # this will be an array of the MSE score 
    CNN_ad_scoread = np.zeros([CNN_ad,1]) # this will be an array of the AD score
    CNN_ad_nrparams = np.zeros([CNN_ad,1]) # thiss will be an array of the number of parameters
    for w in range(len(filenames_CNN)):
        #check whether the model is trained with AD loss or MSE loss
        if '_MSE' in filenames_CNN[w]:
            CNN_mse_names.append(filenames_CNN[w])
            CNN_mse_scoremse[count_CNN_mse] = mean_mse_CNN[w]
            CNN_mse_scoread[count_CNN_mse] = mean_ad_CNN[w]
            modelname_temp = filenames_CNN[w]
            # get nr of model parameters
            for z in range(len(modelsanalyzed)):
                if modelsanalyzed[z].value == modelname_temp[:-16]:
                    CNN_mse_nrparams[count_CNN_mse] = models_nrparams[z].value
                    break
            count_CNN_mse = count_CNN_mse + 1
        elif '_AD' in filenames_CNN[w]:
            CNN_ad_names.append(filenames_CNN[w])
            CNN_ad_scoremse[count_CNN_ad] = mean_mse_CNN[w]
            CNN_ad_scoread[count_CNN_ad] = mean_ad_CNN[w]
            modelname_temp = filenames_CNN[w]
            # get nr of model parameters
            for z in range(len(modelsanalyzed)):
                if modelsanalyzed[z].value == modelname_temp[:-16]:
                    CNN_ad_nrparams[count_CNN_ad] = models_nrparams[z].value
                    break
            count_CNN_ad = count_CNN_ad + 1
    
    # add everything to a dictionary
    CNN_dict_avgperf = dict(CNN_mse_names = CNN_mse_names, CNN_mse_scoremse = CNN_mse_scoremse, CNN_mse_scoread = CNN_mse_scoread, CNN_mse_nrparams = CNN_mse_nrparams, \
                    CNN_ad_names = CNN_ad_names, CNN_ad_scoremse = CNN_ad_scoremse, CNN_ad_scoread = CNN_ad_scoread, CNN_ad_nrparams = CNN_ad_nrparams)
    
    return CNN_dict_avgperf