# script to analyze model performance

# to clear all variables: type in ipython console %reset

#------------------------------------------------------------------------------
# Specifications
#------------------------------------------------------------------------------

# set directories
dirfiles = 'workspace/notebooks/histories'
dirscripts = '/workspace/notebooks/scripts'
excelfile = '/workspace/notebooks/scripts/testfile.xlsx'

# import required packages and libraries
import os
import numpy as np
import pandas
import openpyxl

# create list of file names
os.chdir(dirfiles)
filenames = [x for x in os.listdir(dirfiles) if x.endswith(".csv")]

#-----------------------------------------------------------------------------------
# for each model, retrieve relevant information from history file
#-----------------------------------------------------------------------------------
for x in range(len(filenames)):
    # get current model name
    modelname_temp = filenames[x]
    modelname_temp_short = modelname_temp[8:-4] # remove the extension and 'history'
    # open excel file
    wb = openpyxl.load_workbook(excelfile) # open excel file
    ws = wb['RawData'] # select active sheet
    # check whether this model has already been processed and added to the workbook before
    modelsanalyzed = ws['A'] # get names of analyzed models from first column
    vectcheck = np.zeros(len(modelsanalyzed))
    for y in range(len(modelsanalyzed)):
        if modelsanalyzed[y].value == modelname_temp_short:
            vectcheck[y] = 1
    if sum(vectcheck) == 0: # only continue if the model has not been analyzed before      
        # load history of model
        hist = pandas.read_csv(modelname_temp)
        colnames = list(hist.columns)
        # retrieve number of epochs
        nrepochs = hist.iloc[-1,0] # you need to use iloc for indexing
        # retrieve information about training
        stop_loss = hist.loc[nrepochs-10,'loss'] # patience for training was set to 10, so get the values of lowest point after which no improvement occured
        stop_val_loss = hist.loc[nrepochs-10,'val_loss']
        if 'val_mse' in colnames:
            stop_val_mse = hist.loc[nrepochs-10,'val_mse']
        elif 'val_mean_squared_error' in colnames:
            stop_val_mse = hist.loc[nrepochs-10,'val_mean_squared_error']
        stop_val_cos_2d_angular = hist.loc[nrepochs-10,'val_cos_distmet_2D_angular']
        # open excel file
        wb = openpyxl.load_workbook(excelfile) # open excel file
        ws = wb['RawData'] # select active sheet
        # find first open row
        rownr = ws.max_row+1
        ws.cell(column = 1, row = rownr, value = modelname_temp_short) # write to cell
        ws.cell(column = 2, row = rownr, value = nrepochs) # write to cell
        ws.cell(column = 3, row = rownr, value = stop_loss) # write to cell
        ws.cell(column = 4, row = rownr, value = stop_val_loss) # write to cell
        ws.cell(column = 5, row = rownr, value = stop_val_mse)
        ws.cell(column = 6, row = rownr, value = stop_val_cos_2d_angular)
        wb.save(excelfile) # save excel file

