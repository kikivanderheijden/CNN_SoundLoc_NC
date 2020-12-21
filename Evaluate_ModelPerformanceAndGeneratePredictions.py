# script to analyze model performance

# to clear all variables: type in ipython console %reset

#------------------------------------------------------------------------------
# Specifications
#------------------------------------------------------------------------------

# set directories
dirfiles = '/workspace/notebooks/histories/'
dirscripts = '/workspace/notebooks/scripts'
dirsounds = '/workspace/sounds_small_npy/eval'

# import required packages and libraries
from tensorflow.keras.models import load_model
import tensorflow.keras.backend as K
import os
import numpy as np
import pandas
import pickle
import openpyxl
import matplotlib.pyplot as plt


os.chdir(dirscripts)
from CustLoss_MSE import cust_mean_squared_error
from ModelPredictions import generate_model_predictions
from CustLoss_cosine_distance_angular import cos_dist_2D_angular
from CustMet_cosine_distance_angular import cos_distmet_2D_angular
from CustLoss_Combined_CosineAngular_MSE_weighed import cos_dist_angular_and_mse_weighed # note that in this loss function, the axis of the MSE is set to 1

# create list of file names
os.chdir(dirfiles)
filenames = [x for x in os.listdir(dirfiles) if x.endswith(".csv")]

#-----------------------------------------------------------------------------------
# for each model, retrieve relevant information from history file
#-----------------------------------------------------------------------------------
for x in range(len(filenames)):
    # get current model name
    modelname_temp = filenames[x]
    modelname_temp_short = modelname_temp[8:-4] # remove the extension and 'history'
    # open excel file
    wb = openpyxl.load_workbook(dirscripts+"/Overview_ModelPerformance.xlsx") # open excel file
    ws = wb['RawData'] # select active sheet
    # check whether this model has already been processed and added to the workbook before
    modelsanalyzed = ws['A'] # get names of analyzed models from first column
    vectcheck = np.zeros(len(modelsanalyzed))
    for y in range(len(modelsanalyzed)):
        if modelsanalyzed[y].value == modelname_temp_short:
            vectcheck[y] = 1
    if sum(vectcheck) == 0: # only continue if the model has not been analyzed before      
        # load history of model
        hist = pandas.read_csv(modelname_temp)
        colnames = list(hist.columns)
        # retrieve number of epochs
        nrepochs = hist.iloc[-1,0] # you need to use iloc for indexing
        # retrieve information about training
        stop_loss = hist.loc[nrepochs-10,'loss'] # patience for training was set to 10, so get the values of lowest point after which no improvement occured
        stop_val_loss = hist.loc[nrepochs-10,'val_loss']
        if 'val_mse' in colnames:
            stop_val_mse = hist.loc[nrepochs-10,'val_mse']
        elif 'val_mean_squared_error' in colnames:
            stop_val_mse = hist.loc[nrepochs-10,'val_mean_squared_error']
        stop_val_cos_2d_angular = hist.loc[nrepochs-10,'val_cos_distmet_2D_angular']
        # open excel file
        wb = openpyxl.load_workbook(dirscripts+"/Overview_ModelPerformance.xlsx") # open excel file
        ws = wb['RawData'] # select active sheet
        # find first open row
        rownr = ws.max_row+1
        ws.cell(column = 1, row = rownr, value = modelname_temp_short) # write to cell
        ws.cell(column = 2, row = rownr, value = nrepochs) # write to cell
        ws.cell(column = 3, row = rownr, value = stop_loss) # write to cell
        ws.cell(column = 4, row = rownr, value = stop_val_loss) # write to cell
        ws.cell(column = 5, row = rownr, value = stop_val_mse)
        ws.cell(column = 6, row = rownr, value = stop_val_cos_2d_angular)
        wb.save(dirscripts+"/Overview_ModelPerformance.xlsx") # save excel file

#-----------------------------------------------------------------------------------    
# load data for evaluation and prediction
#-----------------------------------------------------------------------------------        
## model parameters for evaluation
sizebatches = 64

# load the data
an_l_val = np.load(dirsounds+"/an_l_eval_sounds.npy")
an_r_val = np.load(dirsounds+"/an_r_eval_sounds.npy") 
labels_val =  np.load(dirsounds+"/labels_eval_sounds.npy")
names_val = pickle.load(open(dirsounds+'/listfilenames_evall_sounds.p','rb'))
# prepare data for model evaluation
X_test = [an_l_val, an_r_val]
Y_test = labels_val


#-----------------------------------------------------------------------------------    
# for each model, evaluate performance on evaluation data
#-----------------------------------------------------------------------------------        
# check whether the evaluation of this model has already been performed
for x in range(len(filenames)):
    # get current model name
    modelname_temp = filenames[x]
    modelname_temp_short = modelname_temp[8:-4] # remove the extension and 'history'
    # open excel file
    wb = openpyxl.load_workbook(dirscripts+"/Overview_ModelPerformance.xlsx") # open excel file
    ws = wb['RawData'] # select sheet
    # check whether this model has already been processed and added to the workbook before
    modelsanalyzed = ws['A'] # get names of analyzed models from first column
    vectcheck1 = np.zeros(len(modelsanalyzed))
    for y in range(len(modelsanalyzed)):
        if modelsanalyzed[y].value == modelname_temp_short:
            vectcheck1[y] = 1
    if sum(vectcheck1) > 0: # check whether the model has also been evaluated before   
        testval = ws.cell(column = 7, row = vectcheck1.tolist().index(1)+1).value # add one because excel starts from number 1 while pythons starts from zero
    if sum(vectcheck1) == 0 or testval is None: # only continue if this model hasn't been evaluated yet    
        # evaluate the model
        model = load_model(dirfiles+'/'+modelname_temp_short+'_final.h5', custom_objects={"cust_mean_squared_error": cust_mean_squared_error, "cos_dist_2D_angular": cos_dist_2D_angular, "cos_distmet_2D_angular": cos_distmet_2D_angular, "cos_dist_angular_and_mse_weighed": cos_dist_angular_and_mse_weighed})
        score = model.evaluate(X_test, Y_test, verbose=1)
        # write scores to excel file, if the model is evaluated you should add it to that row, otherwise new row with model name + scores
        if sum(vectcheck1) == 0:
            # find first open row
            rownr = ws.max_row+1
            ws.cell(column = 1, row = rownr, value = modelname_temp_short) # write to cell
            ws.cell(column = 7, row = rownr, value = score[2]) # write to cell
            ws.cell(column = 8, row = rownr, value = score[3]) # write to cell
        elif sum(vectcheck1) > 0:
            ws.cell(column = 7, row =  vectcheck1.tolist().index(1)+1, value = score[2]) # write to cell
            ws.cell(column = 8, row =  vectcheck1.tolist().index(1)+1, value = score[3]) # write to cell            
        wb.save(dirscripts+"/Overview_ModelPerformance.xlsx") # save excel file


#----------------------------------------------------------------------------------
# for each model, retrieve number of parameters
#----------------------------------------------------------------------------------
# check whether the evaluation of this model has already been performed
for x in range(len(filenames)):
    # get current model name
    modelname_temp = filenames[x]
    modelname_temp_short = modelname_temp[8:-4] # remove the extension and 'history'
    # open excel file
    wb = openpyxl.load_workbook(dirscripts+"/Overview_ModelPerformance.xlsx") # open excel file
    ws = wb['RawData'] # select active sheet
    # check whether this model has already been processed and added to the workbook before
    modelsanalyzed = ws['A'] # get names of analyzed models from first column
    vectcheck1 = np.zeros(len(modelsanalyzed))
    for y in range(len(modelsanalyzed)):
        if modelsanalyzed[y].value == modelname_temp_short:
            vectcheck1[y] = 1
    if sum(vectcheck1) > 0: # check whether the model parameters have already been retrieved  
        testval = ws.cell(column = 9, row = vectcheck1.tolist().index(1)+1).value # add one because excel starts from number 1 while pythons starts from zero
    if sum(vectcheck1) == 0 or testval is None: # only continue if the model parameters have not yet been retrieved  
        model = load_model(dirfiles+'/'+modelname_temp_short+'_final.h5', custom_objects={"cust_mean_squared_error": cust_mean_squared_error, "cos_dist_2D_angular": cos_dist_2D_angular, "cos_distmet_2D_angular": cos_distmet_2D_angular, "cos_dist_angular_and_mse_weighed": cos_dist_angular_and_mse_weighed})
        trainable_count = np.sum([K.count_params(w) for w in model.trainable_weights])
        non_trainable_count = np.sum([K.count_params(w) for w in model.non_trainable_weights])
        # write scores to excel file, if the model is evaluated you should add it to that row, otherwise new row with model name + scores
        if sum(vectcheck1) == 0:
            # find first open row
            rownr = ws.max_row+1
            ws.cell(column = 1, row = rownr, value = modelname_temp_short) # write to cell
            ws.cell(column = 9, row = rownr, value = trainable_count + non_trainable_count) # write to cell
        elif sum(vectcheck1) > 0:
            ws.cell(column = 9, row =  vectcheck1.tolist().index(1)+1, value = trainable_count + non_trainable_count) # write to cell
        wb.save(dirscripts+"/Overview_ModelPerformance.xlsx") # save excel file
        
        
#-----------------------------------------------------------------------------------
# for each model, generate model predictions
#-----------------------------------------------------------------------------------

# create list of names of prediction files
os.chdir(dirfiles)
filenames = [x for x in os.listdir(dirfiles) if x.endswith(".csv")]

for x in range(len(filenames)):
    # check whether existing predictions were made
    existingpreds = [x for x in os.listdir(dirfiles) if x.endswith("predictions.npy")] # predictions are saved as npy files
    # get current model name
    modelname_temp = filenames[x]
    modelname_temp_short = modelname_temp[8:-4] # remove the extension and 'history'
    if len(existingpreds) != 0:
        vectcheck2 = np.zeros(len(existingpreds)) # create a vector for checking
        for y in range(len(existingpreds)):
            if modelname_temp_short in existingpreds[y]:
                vectcheck2[y] = 1
        if sum(vectcheck2) == 0: # only execute predictions if they haven't been done before
            model = load_model(dirfiles+'/'+modelname_temp_short+'_final.h5', custom_objects={"cust_mean_squared_error": cust_mean_squared_error, "cos_dist_2D_angular": cos_dist_2D_angular, "cos_distmet_2D_angular": cos_distmet_2D_angular, "cos_dist_angular_and_mse_weighed": cos_dist_angular_and_mse_weighed})
            predictions = generate_model_predictions(model, X_test, modelname_temp_short, dirfiles, sizebatches)
    elif len(existingpreds) == 0: # in case there are no prediction files yet
        model = load_model(dirfiles+'/'+modelname_temp_short+'_final.h5', custom_objects={"cust_mean_squared_error": cust_mean_squared_error, "cos_dist_2D_angular": cos_dist_2D_angular, "cos_distmet_2D_angular": cos_distmet_2D_angular, "cos_dist_angular_and_mse_weighed": cos_dist_angular_and_mse_weighed})
        predictions = generate_model_predictions(model, X_test, modelname_temp_short, dirfiles, sizebatches)
           

#-----------------------------------------------------------------------------------
# for each model, create figure of training history
#-----------------------------------------------------------------------------------
# creating a plot of the loss (i.e. model performance)

# create list of histories
os.chdir(dirfiles)
filenames = [x for x in os.listdir(dirfiles) if x.endswith(".csv")]
filenames_check = [x for x in os.listdir(dirfiles) if x.endswith(".png")]


for x in range(len(filenames)):
    # check whether figure was already created
    modelname_temp = filenames[x]
    modelname_temp_short = modelname_temp[8:-4] # remove the extension and 'history'
    if not modelname_temp_short+'_traininghistory.png' in filenames_check: # check whether the figure already exists and only continue if it doesn't 
       # read the file with pandas
       hist = pandas.read_csv(dirfiles+'/'+modelname_temp)
       # you have to make a distinction for models with AD loss and with MSE loss
       if '_AD' in modelname_temp:
           plt.figure()
           plt.plot(hist.loss*180) # multiply with 180 to make it angular
           plt.plot(hist.val_loss*180)
           if len(modelname_temp_short) > 50:
               plt.title("Training History \n "+modelname_temp_short[0:40]+"\n"+modelname_temp_short[41:])   
           elif len(modelname_temp_short) < 50:
               plt.title("Training History \n "+modelname_temp_short)
           plt.ylim(0, 50)
           plt.ylabel("AD loss (degrees)")
           plt.xlabel("epoch")
           plt.legend(["Train", "Test"], loc="upper right")
       elif 'MSE' in modelname_temp: 
           plt.figure()
           plt.plot(hist.loss) 
           plt.plot(hist.val_loss)
           if len(modelname_temp_short) > 50:
               plt.title("Training History \n "+modelname_temp_short[0:40]+"\n"+modelname_temp_short[41:])   
           elif len(modelname_temp_short) < 50:
               plt.title("Training History \n "+modelname_temp_short)
           plt.ylim(0,0.25)
           plt.ylabel("MSE loss")
           plt.xlabel("epoch")
           plt.legend(["Train", "Test"], loc="upper right")
       plt.savefig(dirfiles+'/'+modelname_temp_short+'_traininghistory.png')
       plt.close('all')
    
